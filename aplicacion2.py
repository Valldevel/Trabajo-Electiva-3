from flask import Flask, request, render_template_string, jsonify
from tensorflow.keras.models import load_model
import numpy as np
from PIL import Image
import io
import base64
import openpyxl
from openpyxl import Workbook
from datetime import datetime

app = Flask(__name__)

# Lista de clases (etiquetas) de tu modelo
class_names = [
    'TANGARA DE LENTEJUELAS', 'TANGARA MATORRALERA', 'TANGARA NUQUIRRUFA', 'TANGARA PINTOJA', 'PINTOJA PALMERA',
    'TANGARA AZULGRIS', 'TANGARA AZULINEGRA', 'TANGARA CABECIAZUL', 'TANGARA CAPUCHA DORADA', 'TANGARA CORONINEGRA',
    'CHIPE AMARILLO FINAL', 'CHIPE CABEZA NEGRA FINAL', 'CHIPE CASTAÑO FINAL', 'CHIPE DORSO VERDE FINAL',
    'CHIPE FLANCOS CASTAÑOS FINAL', 'CHIPE GARGANTA AMARILLA FINAL', 'CHIPE GARGANTA NARANJA FINAL',
    'CHIPE GORRA CANELA SUREÑO FINAL', 'TROPICAL PARULA FINAL'
]

# Descripciones de las clases
class_descriptions = {
    'TANGARA DE LENTEJUELAS': "La tangara de lentejuelas se caracteriza por su vibrante coloración verde y azul, con reflejos metálicos que imitan lentejuelas.",
    'TANGARA MATORRALERA': "Esta especie habita los matorrales y tiene un plumaje colorido que va desde el verde hasta el rojo y amarillo.",
    'TANGARA NUQUIRRUFA': "Una tangara de tonos predominantes en verde y amarillos, conocida por su canto melodioso.",
    'TANGARA PINTOJA': "La tangara pintoja es una especie que destaca por su colorido plumaje en tonos marrones y anaranjados.",
    'PINTOJA PALMERA': "Esta especie habita las palmas de América Central, y su plumaje presenta una mezcla de tonos marrones y verdes.",
    'TANGARA AZULGRIS': "Con su distintivo color azul grisáceo, esta tangara se encuentra en las selvas tropicales de Sudamérica.",
    'TANGARA AZULINEGRA': "Una especie con colores que van desde el azul eléctrico hasta el negro, visible en zonas de montaña.",
    'TANGARA CABECIAZUL': "La tangara cabeciazul tiene una característica coloración azul en su cabeza y pecho, con el resto de su cuerpo en tonos verdes.",
    'TANGARA CAPUCHA DORADA': "Con su característica capucha dorada, esta tangara es un espectáculo visual en las selvas tropicales.",
    'TANGARA CORONINEGRA': "Su característica es su corona negra y plumaje colorido en tonos rojos y amarillos.",
    'CHIPE AMARILLO FINAL': "Un pequeño chip de color amarillo brillante que habita áreas abiertas y bordes de bosque.",
    'CHIPE CABEZA NEGRA FINAL': "Este chip tiene una cabeza negra contrastante con el resto de su cuerpo de colores más apagados.",
    'CHIPE CASTAÑO FINAL': "Un chip de color castaño, con una excelente capacidad para camuflarse en su hábitat natural.",
    'CHIPE DORSO VERDE FINAL': "Con un característico dorso verde, este chip es común en áreas boscosas.",
    'CHIPE FLANCOS CASTAÑOS FINAL': "Este chip tiene flancos castaños y una notable capacidad para moverse rápidamente entre los arbustos.",
    'CHIPE GARGANTA AMARILLA FINAL': "Con una garganta de color amarillo brillante, este chip es fácil de identificar entre la vegetación.",
    'CHIPE GARGANTA NARANJA FINAL': "Similar al anterior, pero con una garganta de color naranja intenso.",
    'CHIPE GORRA CANELA SUREÑO FINAL': "Este chip se destaca por su gorra de color canela, especialmente visible en los climas sureños.",
    'TROPICAL PARULA FINAL': "Un pequeño y vibrante pájaro de los trópicos, conocido por sus colores brillantes y su comportamiento activo."
}

# Cargar modelo entrenado
model = load_model('models/model_VGG16_v.keras')

# Función para guardar avistamientos en el Excel
def guardar_avistamiento(lat, lng, prediction):
    archivo = "avistamientos.xlsx"
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        wb = openpyxl.load_workbook(archivo)
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Latitud", "Longitud", "Predicción", "Fecha y Hora"])
    sheet.append([lat, lng, prediction, now])
    wb.save(archivo)

# Obtener los avistamientos registrados desde el archivo Excel
def obtener_avistamientos():
    archivo = "avistamientos.xlsx"
    avistamientos = []
    try:
        wb = openpyxl.load_workbook(archivo)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            avistamientos.append({
                'lat': row[0],
                'lng': row[1],
                'prediction': row[2],
                'fecha_hora': row[3]
            })
    except FileNotFoundError:
        print(f"Error: El archivo {archivo} no fue encontrado.")
    except Exception as e:
        print(f"Error al leer el archivo: {e}")
    
    return avistamientos

# HTML con JavaScript para pasar predicción a JS
html_template = '''
<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Clasificador de Aves</title>
  <style>
    body {
      font-family: Arial, sans-serif;
      background-color: #f0f8ff;
      margin: 0;
      padding: 0;
      color: #333;
    }

    h1, h2 {
      color: #007BFF;
      text-align: center;
    }

    .container {
      background: white;
      padding: 20px;
      border-radius: 8px;
      width: 80%;
      max-width: 900px;
      margin: 20px auto;
      box-shadow: 0 4px 8px rgba(0, 0, 0, 0.1);
    }

    .container input[type="file"],
    .container input[type="number"],
    .container button {
      padding: 10px;
      margin: 10px 0;
      width: 100%;
      box-sizing: border-box;
      font-size: 16px;
      border-radius: 5px;
      border: 1px solid #ccc;
    }

    .container input[type="submit"] {
      background-color: #28a745;
      color: white;
      cursor: pointer;
      border: none;
    }

    .container input[type="submit"]:hover {
      background-color: #218838;
    }

    .map-container {
      width: 100%;
      height: 400px;
      margin-top: 20px;
    }

    #map {
      width: 100%;
      height: 100%;
      border-radius: 8px;
    }

    table {
      width: 100%;
      margin-top: 20px;
      border-collapse: collapse;
      background-color: #fafafa;
      border-radius: 8px;
      box-shadow: 0 2px 5px rgba(0, 0, 0, 0.1);
    }

    th, td {
      padding: 12px;
      text-align: center;
      border-bottom: 1px solid #ddd;
    }

    th {
      background-color: #007BFF;
      color: white;
    }

    tr:hover {
      background-color: #f1f1f1;
    }

    .button {
      background-color: #007BFF;
      color: white;
      padding: 10px 20px;
      border: none;
      border-radius: 5px;
      cursor: pointer;
      margin-top: 20px;
    }

    .button:hover {
      background-color: #0056b3;
    }

    .alert {
      color: green;
      text-align: center;
      font-size: 18px;
      margin-top: 20px;
    }

    img {
      border-radius: 8px;
      box-shadow: 0 4px 6px rgba(0, 0, 0, 0.2);
      item-align: center;
      margin-top: 20px;
    }

    .form-group {
      display: flex;
      flex-direction: column;
      align-items: center;
    }
  </style>
</head>
<body>
  <div class="container">
    <h1>Clasificador de Aves</h1>
    <form method="POST" enctype="multipart/form-data">
      <div class="form-group">
        <input type="file" name="file" accept="image/*" required />
        <input type="submit" value="Clasificar" />
      </div>
    </form>

    {% if prediction %}
      <h2>Predicción: {{ prediction }}</h2>
      <p>{{ description }}</p>
      <img src="data:image/jpeg;base64,{{ image_data }}" width="300" />
    {% endif %}
  </div>

  <div class="container">
    <h2>Registrar Avistamiento</h2>
    <div class="form-group">
      <input type="number" id="lat" placeholder="Latitud" step="any" required />
      <input type="number" id="lng" placeholder="Longitud" step="any" required />
      <button class="button" onclick="addBirdSighting()">Guardar Avistamiento</button>
    </div>
  </div>

  <div class="map-container">
    <div id="map"></div>
  </div>

  <div class="container">
    <h2>Avistamientos Registrados</h2>
    <table id="sighting-table">
      <thead>
        <tr>
          <th>Latitud</th>
          <th>Longitud</th>
          <th>Predicción</th>
          <th>Fecha y Hora</th>
        </tr>
      </thead>
      <tbody>
        <!-- Los avistamientos se cargarán aquí -->
      </tbody>
    </table>
  </div>

  {% if prediction %}
    <script>setPrediction('{{ prediction }}');</script>
  {% endif %}

  <script>
    let map;
    let avistamientos = [];

    function initMap() {
      map = new google.maps.Map(document.getElementById("map"), {
        center: { lat: 0, lng: 0 },
        zoom: 2
      });

      // Cargar los avistamientos desde el servidor
      fetch('/get_avistamientos')
        .then(response => response.json())
        .then(data => {
          avistamientos = data;
          avistamientos.forEach(avistamiento => {
            const { lat, lng, prediction } = avistamiento;
            new google.maps.Marker({
              position: { lat, lng },
              map: map,
              title: prediction
            });
          });
        });
    }

    function addBirdSighting() {
      const lat = parseFloat(document.getElementById('lat').value);
      const lng = parseFloat(document.getElementById('lng').value);
      const prediction = '{{ prediction }}';

      if (lat && lng) {
        fetch('/guardar_avistamiento', {
          method: 'POST',
          headers: {
            'Content-Type': 'application/json'
          },
          body: JSON.stringify({ lat, lng, prediction })
        })
        .then(response => response.json())
        .then(data => alert('Avistamiento guardado exitosamente'))
        .catch(error => console.error('Error al guardar avistamiento:', error));
      } else {
        alert('Por favor, ingrese una latitud y longitud válidas.');
      }
    }

    // Función para cargar y mostrar los avistamientos en la tabla
    function loadSightingTable() {
      fetch('/get_avistamientos')
        .then(response => response.json())
        .then(data => {
          const table = document.getElementById('sighting-table');
          table.innerHTML = '';  // Limpiar tabla
          const header = document.createElement('tr');
          header.innerHTML = '<th>Latitud</th><th>Longitud</th><th>Predicción</th><th>Fecha y Hora</th>';
          table.appendChild(header);

          data.forEach(avistamiento => {
            const row = document.createElement('tr');
            row.innerHTML = `<td>${avistamiento.lat}</td><td>${avistamiento.lng}</td><td>${avistamiento.prediction}</td><td>${avistamiento.fecha_hora}</td>`;
            table.appendChild(row);
          });
        })
        .catch(error => console.error('Error al cargar los avistamientos:', error));
    }

    // Cargar la tabla al cargar la página
    window.onload = loadSightingTable;
  </script>

  <!-- Aquí agregas tu clave API de Google Maps -->
  <script async defer
    src="https://maps.googleapis.com/maps/api/js?key=AIzaSyCZR_MdAc09QAW0nWJvlCdcwIx_CQQoM2Y&callback=initMap">
  </script>
</body>
</html>



'''

# Ruta principal para clasificar la imagen
@app.route('/', methods=['GET', 'POST'])
def classify_image():
    prediction = None
    description = None
    image_data = None

    if request.method == 'POST':
        file = request.files['file']
        if file:
            img = Image.open(file.stream).convert('RGB').resize((224, 224))
            img_array = np.expand_dims(np.array(img) / 255.0, axis=0)
            preds = model.predict(img_array)
            idx = np.argmax(preds, axis=1)[0]
            prediction = class_names[idx]
            description = class_descriptions.get(prediction, "Descripción no disponible.")
            buffered = io.BytesIO()
            img.save(buffered, format="JPEG")
            image_data = base64.b64encode(buffered.getvalue()).decode('utf-8')

    return render_template_string(
        html_template,
        prediction=prediction,
        description=description,
        image_data=image_data
    )

# Ruta para guardar el avistamiento
@app.route('/guardar_avistamiento', methods=['POST'])
def guardar_avistamiento_route():
    data = request.get_json()
    lat = data.get('lat')
    lng = data.get('lng')
    prediction = data.get('prediction', "Sin predicción")
    guardar_avistamiento(lat, lng, prediction)
    return jsonify({"message": "Avistamiento guardado exitosamente"}), 200

# Ruta para obtener los avistamientos desde el archivo Excel
@app.route('/get_avistamientos')
def get_avistamientos():
    avistamientos = obtener_avistamientos()
    return jsonify(avistamientos)

if __name__ == '__main__':
    app.run(debug=True)
