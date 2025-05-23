import streamlit as st
import numpy as np
from PIL import Image
import base64
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from tensorflow.keras.models import load_model
import io
import pandas as pd
import streamlit.components.v1 as components
import gdown

# Lista de clases (etiquetas) de tu modelo
class_names = [
    'TANGARA DE LENTEJUELAS', 'TANGARA MATORRALERA', 'TANGARA NUQUIRRUFA', 'TANGARA PINTOJA', 'PINTOJA PALMERA',
    'TANGARA AZULGRIS', 'TANGARA AZULINEGRA', 'TANGARA CABECIAZUL', 'TANGARA CAPUCHA DORADA', 'TANGARA CORONINEGRA',
    'CHIPE AMARILLO FINAL', 'CHIPE CABEZA NEGRA FINAL', 'CHIPE CASTAÑO FINAL', 'CHIPE DORSO VERDE FINAL',
    'CHIPE FLANCOS CASTAÑOS FINAL', 'CHIPE GARGANTA AMARILLA FINAL', 'CHIPE GARGANTA NARANJA FINAL',
    'CHIPE GORRA CANELA SUREÑO FINAL', 'TROPICAL PARULA FINAL'
]

# Descripciones de las clases
class_descriptions = {
    'TANGARA DE LENTEJUELAS': "La tangara de lentejuelas se caracteriza por su vibrante coloración verde y azul, con reflejos metálicos que imitan lentejuelas.",
    'TANGARA MATORRALERA': "Esta especie habita los matorrales y tiene un plumaje colorido que va desde el verde hasta el rojo y amarillo.",
    'TANGARA NUQUIRRUFA': "Una tangara de tonos predominantes en verde y amarillos, conocida por su canto melodioso.",
    'TANGARA PINTOJA': "La tangara pintoja es una especie que destaca por su colorido plumaje en tonos marrones y anaranjados.",
    'PINTOJA PALMERA': "Esta especie habita las palmas de América Central, y su plumaje presenta una mezcla de tonos marrones y verdes.",
    'TANGARA AZULGRIS': "Con su distintivo color azul grisáceo, esta tangara se encuentra en las selvas tropicales de Sudamérica.",
    'TANGARA AZULINEGRA': "Una especie con colores que van desde el azul eléctrico hasta el negro, visible en zonas de montaña.",
    'TANGARA CABECIAZUL': "La tangara cabeciazul tiene una característica coloración azul en su cabeza y pecho, con el resto de su cuerpo en tonos verdes.",
    'TANGARA CAPUCHA DORADA': "Con su característica capucha dorada, esta tangara es un espectáculo visual en las selvas tropicales.",
    'TANGARA CORONINEGRA': "Su característica es su corona negra y plumaje colorido en tonos rojos y amarillos.",
    'CHIPE AMARILLO FINAL': "Un pequeño chip de color amarillo brillante que habita áreas abiertas y bordes de bosque.",
    'CHIPE CABEZA NEGRA FINAL': "Este chip tiene una cabeza negra contrastante con el resto de su cuerpo de colores más apagados.",
    'CHIPE CASTAÑO FINAL': "Un chip de color castaño, con una excelente capacidad para camuflarse en su hábitat natural.",
    'CHIPE DORSO VERDE FINAL': "Con un característico dorso verde, este chip es común en áreas boscosas.",
    'CHIPE FLANCOS CASTAÑOS FINAL': "Este chip tiene flancos castaños y una notable capacidad para moverse rápidamente entre los arbustos.",
    'CHIPE GARGANTA AMARILLA FINAL': "Con una garganta de color amarillo brillante, este chip es fácil de identificar entre la vegetación.",
    'CHIPE GARGANTA NARANJA FINAL': "Similar al anterior, pero con una garganta de color naranja intenso.",
    'CHIPE GORRA CANELA SUREÑO FINAL': "Este chip se destaca por su gorra de color canela, especialmente visible en los climas sureños.",
    'TROPICAL PARULA FINAL': "Un pequeño y vibrante pájaro de los trópicos, conocido por sus colores brillantes y su comportamiento activo."
}

# Cargar modelo desde Google Drive
file_id = "1rh2AvU4O0aboTqIN_BLh_u1D5J868Aln"
url = f"https://drive.google.com/uc?id={file_id}"
model_path = "best_model.keras"
gdown.download(url, model_path, quiet=False)
model = load_model(model_path)
print("✅ Modelo cargado correctamente desde Google Drive")

# Función para guardar avistamiento en Excel
def guardar_avistamiento(lat, lng, prediction):
    archivo = "avistamientos.xlsx"
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        wb = openpyxl.load_workbook(archivo)
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Latitud", "Longitud", "Predicción", "Fecha y Hora"])
    sheet.append([lat, lng, prediction, now])
    wb.save(archivo)

# Función para obtener avistamientos
def obtener_avistamientos():
    archivo = "avistamientos.xlsx"
    avistamientos = []
    try:
        wb = openpyxl.load_workbook(archivo)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            avistamientos.append({
                'lat': row[0],
                'lng': row[1],
                'prediction': row[2],
                'fecha_hora': row[3]
            })
    except FileNotFoundError:
        print(f"Error: El archivo {archivo} no fue encontrado.")
    except Exception as e:
        print(f"Error al leer el archivo: {e}")
    return avistamientos

# Interfaz Streamlit
st.title('Avistamiento de Aves')

# Subir imagen
uploaded_file = st.file_uploader("Sube una imagen de un ave", type=["jpg", "jpeg", "png"])

prediction = None
description = None

if uploaded_file is not None:
    img = Image.open(uploaded_file).convert('RGB').resize((224, 224))
    img_array = np.expand_dims(np.array(img) / 255.0, axis=0)
    preds = model.predict(img_array)
    idx = np.argmax(preds, axis=1)[0]
    prediction = class_names[idx]
    description = class_descriptions.get(prediction, "Descripción no disponible.")

    st.image(img, caption="Imagen clasificada", use_column_width=True)
    st.subheader(f"Predicción: {prediction}")
    st.write(description)

# Registro de coordenadas
lat = st.number_input('Latitud', format="%.6f")
lng = st.number_input('Longitud', format="%.6f")

if st.button('Guardar Avistamiento'):
    if lat and lng and prediction:
        guardar_avistamiento(lat, lng, prediction)
        st.success('Avistamiento guardado exitosamente!')
    else:
        st.error('Por favor, asegúrate de haber cargado una imagen y de ingresar una latitud y longitud válidas.')

# Mostrar tabla y mapa de avistamientos
st.subheader('Avistamientos Registrados')
avistamientos = obtener_avistamientos()

if avistamientos:
    df_avistamientos = pd.DataFrame(avistamientos)
    st.dataframe(df_avistamientos)  # Tabla interactiva

    # Mapa de Google Maps con múltiples marcadores
    st.subheader('Mapa de Google Maps')

    markers_js = ""
    for _, row in df_avistamientos.iterrows():
        markers_js += f"""
        new google.maps.Marker({{
            position: {{ lat: {row['lat']}, lng: {row['lng']} }},
            map: map,
            title: "{row['prediction']}"
        }});"""

    html_code = f"""
    <!DOCTYPE html>
    <html>
      <head>
        <script async defer
          src="https://maps.googleapis.com/maps/api/js?key=AIzaSyCZR_MdAc09QAW0nWJvlCdcwIx_CQQoM2Y&callback=initMap">
        </script>
        <script>
          function initMap() {{
            var map = new google.maps.Map(document.getElementById('map'), {{
              zoom: 6,
              center: {{ lat: {df_avistamientos['lat'].mean()}, lng: {df_avistamientos['lng'].mean()} }}
            }});
            {markers_js}
          }}
        </script>
      </head>
      <body onload="initMap()">
        <div id="map" style="height: 500px; width: 100%;"></div>
      </body>
    </html>
    """

    components.html(html_code, height=600)

    # Detalles adicionales
    for av in avistamientos:
        st.write(f"📍 Lat: {av['lat']}, Lng: {av['lng']} — 🐦 {av['prediction']} — 🕓 {av['fecha_hora']}")
else:
    st.write("No se han registrado avistamientos aún.")
